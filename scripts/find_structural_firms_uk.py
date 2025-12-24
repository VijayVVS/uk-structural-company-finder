import os
import re
import time
import math
import json
import requests
from datetime import datetime, timezone
from urllib.parse import urlparse, urljoin
from openpyxl import Workbook, load_workbook

# ---------- Config ----------
API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")
DAILY_LIMIT = int(os.getenv("DAILY_LIMIT", "25"))
RADIUS_M = int(os.getenv("SEARCH_RADIUS_M", "30000"))
QUERY = os.getenv("SEARCH_QUERY", "civil and structural engineers")

# We hardcode SK3 9AR coords to avoid Geocoding API entirely:
# Stockport / Edgeley approx:
HOME_LAT = float(os.getenv("HOME_LAT", "53.401439"))
HOME_LNG = float(os.getenv("HOME_LNG", "-2.168095"))

if not API_KEY:
    raise SystemExit("Missing GOOGLE_MAPS_API_KEY (GitHub repo secret).")

UA = {"User-Agent": "FBC-CompanyFinder/1.0"}
TIMEOUT = 25

# Only accept role-based emails (avoid personal email harvesting)
GENERIC_EMAIL = re.compile(
    r"\b(careers|recruitment|recruiting|jobs|talent|people|hiring|vacancies|hr)"
    r"@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
    re.IGNORECASE
)

CANDIDATE_PATHS = [
    "/", "/careers", "/careers/", "/jobs", "/jobs/", "/join-us", "/join-us/",
    "/work-with-us", "/work-with-us/", "/contact", "/contact/", "/contact-us", "/contact-us/",
    "/vacancies", "/vacancies/", "/join", "/join/"
]

SEEN_PATH = "data/seen_companies.json"
MASTER_PATH = "data/master_companies.xlsx"
OUT_DAILY_DIR = "out"

# ---------- Helpers ----------
def http_get(url: str, headers=None, params=None):
    r = requests.get(url, headers=headers or UA, params=params, timeout=TIMEOUT)
    r.raise_for_status()
    return r

def http_post(url: str, headers=None, json_body=None):
    r = requests.post(url, headers=headers or UA, json=json_body, timeout=TIMEOUT)
    r.raise_for_status()
    return r

def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371.0
    p1 = math.radians(lat1); p2 = math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dlon/2)**2
    return 2 * R * math.asin(math.sqrt(a))

def normalize_base(website: str):
    if not website:
        return ""
    p = urlparse(website)
    if not p.scheme:
        website = "https://" + website
        p = urlparse(website)
    return f"{p.scheme}://{p.netloc}"

def extract_generic_emails_from_site(base: str):
    found = set()
    checked = []
    for path in CANDIDATE_PATHS:
        url = urljoin(base, path)
        try:
            html = http_get(url).text
        except Exception:
            continue
        checked.append(url)
        for m in GENERIC_EMAIL.finditer(html):
            found.add(m.group(0))
        time.sleep(0.35)
    return sorted(found), checked

# ---------- Seen state ----------
def load_seen():
    os.makedirs(os.path.dirname(SEEN_PATH), exist_ok=True)
    if not os.path.exists(SEEN_PATH):
        return {"seen_place_ids": [], "seen_domains": []}
    with open(SEEN_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def save_seen(seen):
    with open(SEEN_PATH, "w", encoding="utf-8") as f:
        json.dump(seen, f, indent=2)

# ---------- Excel ----------
def write_daily_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily 25"
    headers = [
        "run_date_utc",
        "company_name",
        "distance_km",
        "address",
        "phone",
        "website",
        "generic_recruitment_emails_on_website",
        "pages_checked",
        "google_maps_place_url",
        "place_id",
        "company_domain"
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    wb.save(out_path)

def append_to_master(rows):
    os.makedirs(os.path.dirname(MASTER_PATH), exist_ok=True)
    headers = [
        "run_date_utc",
        "company_name",
        "distance_km",
        "address",
        "phone",
        "website",
        "generic_recruitment_emails_on_website",
        "pages_checked",
        "google_maps_place_url",
        "place_id",
        "company_domain"
    ]
    if not os.path.exists(MASTER_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Master"
        ws.append(headers)
        wb.save(MASTER_PATH)

    wb = load_workbook(MASTER_PATH)
    ws = wb["Master"]
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    wb.save(MASTER_PATH)

# ---------- Places API (New) ----------
PLACES_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"
PLACES_DETAILS_URL = "https://places.googleapis.com/v1/places/"

def places_search_text(text_query: str, lat: float, lng: float, radius_m: int, page_size=20):
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": API_KEY,
        # field mask for search response
        "X-Goog-FieldMask": "places.id,places.displayName,places.formattedAddress,places.location,places.websiteUri,places.internationalPhoneNumber,places.googleMapsUri"
    }
    body = {
        "textQuery": text_query,
        "maxResultCount": page_size,
        "rankPreference": "DISTANCE",
        "locationBias": {
            "circle": {
                "center": {"latitude": lat, "longitude": lng},
                "radius": float(radius_m)
            }
        }
    }
    data = http_post(PLACES_SEARCH_URL, headers=headers, json_body=body).json()
    return data.get("places", [])

def place_details(place_id: str):
    # place_id is like "places/ChIJ...." or sometimes just the ID; handle both
    pid = place_id
    if not pid.startswith("places/"):
        pid = "places/" + pid

    headers = {
        "X-Goog-Api-Key": API_KEY,
        "X-Goog-FieldMask": "id,displayName,formattedAddress,websiteUri,internationalPhoneNumber,googleMapsUri,location"
    }
    url = f"{PLACES_DETAILS_URL}{pid.replace('places/','')}"
    data = http_get(url, headers=headers).json()
    return data

def get_name(place):
    dn = place.get("displayName", {})
    return dn.get("text", "") if isinstance(dn, dict) else ""

def get_latlng(place):
    loc = place.get("location", {})
    return loc.get("latitude"), loc.get("longitude")

def main():
    seen = load_seen()
    seen_place_ids = set(seen.get("seen_place_ids", []))
    seen_domains = set(seen.get("seen_domains", []))

    # Search near home, closest first
    places = places_search_text(QUERY, HOME_LAT, HOME_LNG, RADIUS_M, page_size=60)

    # Enrich with distance and sort
    enriched = []
    for p in places:
        pid = p.get("id")  # Places API (New) returns "id"
        plat, plng = get_latlng(p)
        dist = 9999.0
        if plat is not None and plng is not None:
            dist = haversine_km(HOME_LAT, HOME_LNG, float(plat), float(plng))
        enriched.append((dist, pid, p))
    enriched.sort(key=lambda x: x[0])

    selected_rows = []
    for dist, pid, p in enriched:
        if len(selected_rows) >= DAILY_LIMIT:
            break
        if not pid or pid in seen_place_ids:
            continue

        # Details (to reliably get website/phone)
        try:
            det = place_details(pid)
        except Exception:
            # even if details fails, mark place id seen to avoid repeat loops
            seen_place_ids.add(pid)
            continue

        website = det.get("websiteUri", "") or ""
        base = normalize_base(website)
        domain = urlparse(base).netloc.lower() if base else ""

        # dedupe by domain
        if domain and domain in seen_domains:
            seen_place_ids.add(pid)
            continue

        emails, checked = ([], [])
        if base:
            emails, checked = extract_generic_emails_from_site(base)

        name = get_name(det)
        addr = det.get("formattedAddress", "")
        phone = det.get("internationalPhoneNumber", "")
        maps_url = det.get("googleMapsUri", "")

        row = {
            "run_date_utc": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "company_name": name,
            "distance_km": round(dist, 1) if dist < 9000 else "",
            "address": addr,
            "phone": phone,
            "website": website,
            "generic_recruitment_emails_on_website": " | ".join(emails),
            "pages_checked": " | ".join(checked[:6]),
            "google_maps_place_url": maps_url,
            "place_id": pid,
            "company_domain": domain
        }
        selected_rows.append(row)

        # update seen immediately
        seen_place_ids.add(pid)
        if domain:
            seen_domains.add(domain)

        time.sleep(0.2)

    # persist seen
    seen["seen_place_ids"] = sorted(seen_place_ids)
    seen["seen_domains"] = sorted(seen_domains)
    save_seen(seen)

    # output excel
    os.makedirs(OUT_DAILY_DIR, exist_ok=True)
    daily_name = f"uk_companies_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    daily_path = os.path.join(OUT_DAILY_DIR, daily_name)
    write_daily_excel(selected_rows, daily_path)
    append_to_master(selected_rows)

    print(f"Selected {len(selected_rows)} new companies (limit={DAILY_LIMIT}).")
    print(f"Daily Excel: {daily_path}")
    print(f"Master Excel: {MASTER_PATH}")
    print(f"Seen updated: {SEEN_PATH}")

if __name__ == "__main__":
    main()
